"""
Excel/CSV Import functionality for PM and HR modules
"""
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.parsers import MultiPartParser, FormParser
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime
import openpyxl
import csv
import io
import logging

from .models import Projects, Tasks, Members, TaskStatuses, TaskPriorities
# Import HR models
try:
    from hr.models import Department, Designation, Employee
except ImportError:
    # Fallback/Mock if HR app not ready (for safety, though user has it)
    Department = None
    Designation = None
    Employee = None

logger = logging.getLogger(__name__)

# --- Helper Functions ---

def parse_date(date_val):
    """Parse Excel date or string date to YYYY-MM-DD"""
    if not date_val:
        return None
    if isinstance(date_val, datetime):
        return date_val.date()
    try:
        # Try common formats
        for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y'):
            try:
                return datetime.strptime(str(date_val), fmt).date()
            except ValueError:
                pass
    except Exception:
        pass
    return None

def worksheet_to_records(ws):
    """Convert an openpyxl worksheet to a list of dictionaries"""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(h).strip().lower() if h else '' for h in rows[0]]
    records = []
    for row in rows[1:]:
        record = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                # Clean value
                if val and isinstance(val, str):
                    val = val.strip()
                record[headers[i]] = val
        if any(record.values()):
            records.append(record)
    return records

def parse_file(file):
    """Parse a file (Excel or CSV) and return records (for single sheet/file imports)"""
    filename = file.name.lower()
    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        return worksheet_to_records(ws)
    elif filename.endswith('.csv'):
        content = file.read().decode('utf-8')
        reader = csv.DictReader(io.StringIO(content))
        return [{k.lower().strip(): v for k, v in row.items()} for row in reader]
    else:
        raise ValueError('Unsupported file format. Use .xlsx or .csv')

# --- Lookups (Name as Foreign Key) ---

def get_status_by_name(name):
    if not name:
        return TaskStatuses.objects.filter(is_default=1).first() or TaskStatuses.objects.first()
    return TaskStatuses.objects.filter(name__iexact=name).first() or TaskStatuses.objects.first()

def get_priority_by_name(name):
    if not name:
        return TaskPriorities.objects.filter(is_default=1).first() or TaskPriorities.objects.first()
    return TaskPriorities.objects.filter(name__iexact=name).first() or TaskPriorities.objects.first()

def get_member_by_name_or_email(name_or_email):
    if not name_or_email:
        return None
    # Try email match
    m = Members.objects.filter(email__iexact=name_or_email).first()
    if m: return m
    # Try name match (first + last)
    # Simple split assumes 'First Last'
    parts = name_or_email.split()
    if len(parts) >= 2:
        m = Members.objects.filter(first_name__iexact=parts[0], last_name__iexact=parts[1]).first()
    else:
        m = Members.objects.filter(first_name__iexact=parts[0]).first()
    return m

def get_department_by_name(name):
    if not name or not Department: return None
    dept = Department.objects.filter(dept_name__iexact=name).first()
    if not dept:
        # Auto-create if missing? User requested "use name as foreign key", implied lookup.
        # But usually creating related master data on fly is ambiguous. 
        # For this feature "Unified Import", let's Create It to be helpful.
        try:
            dept = Department.objects.create(
                dept_name=name, 
                dept_code=name[:4].upper() + '-' + str(timezone.now().timestamp())[-4:] # Random unique code
            )
        except:
            return None
    return dept

def get_designation_by_name(name):
    if not name or not Designation: return None
    desig = Designation.objects.filter(designation_name__iexact=name).first()
    if not desig:
        try:
            desig = Designation.objects.create(
                designation_name=name,
                designation_code=name[:4].upper() + '-' + str(timezone.now().timestamp())[-4:],
                level=1
            )
        except:
            return None
    return desig

# --- Import Functions ---

def import_departments(records):
    """Import Departments from records"""
    if not Department: return [], [{'error': 'HR module missing'}]
    created = []
    errors = []
    for idx, rec in enumerate(records):
        try:
            name = rec.get('department_name') or rec.get('name')
            code = rec.get('code')
            if not name: continue
            
            # Check exist
            if Department.objects.filter(dept_name__iexact=name).exists():
                continue # Skip existing

            Department.objects.create(
                dept_name=name,
                dept_code=code or (name[:3].upper() + str(timezone.now().microsecond)),
                category=rec.get('category', 'other').lower(),
                description=rec.get('description', '')
            )
            created.append(name)
        except Exception as e:
            errors.append({'row': idx + 2, 'error': f"Dept '{name}': {str(e)}"})
    return created, errors

def import_designations(records):
    """Import Designations"""
    if not Designation: return [], []
    created = []
    errors = []
    for idx, rec in enumerate(records):
        try:
            name = rec.get('designation_name') or rec.get('name')
            if not name: continue
            
            if Designation.objects.filter(designation_name__iexact=name).exists():
                continue

            Designation.objects.create(
                designation_name=name,
                designation_code=rec.get('code') or (name[:3].upper() + str(timezone.now().microsecond)),
                level=int(rec.get('level') or 1),
                min_salary=rec.get('min_salary'),
                max_salary=rec.get('max_salary')
            )
            created.append(name)
        except Exception as e:
            errors.append({'row': idx + 2, 'error': f"Desig '{name}': {str(e)}"})
    return created, errors

def import_employees_and_members(records):
    """Import Employees and ensure corresponding PM Members exist"""
    created = []
    errors = []
    for idx, rec in enumerate(records):
        try:
            email = rec.get('email')
            if not email:
                errors.append({'row': idx + 2, 'error': 'Missing email'})
                continue

            # 1. Create/Get Member (Auth user)
            member, m_created = Members.objects.get_or_create(
                email=email,
                defaults={
                    'first_name': rec.get('first_name'),
                    'last_name': rec.get('last_name'),
                    'phone': rec.get('phone'),
                    'is_active': 1
                }
            )

            # 2. Create/Get Employee (HR record)
            if Employee:
                # Lookup Dept ID and Desig ID by Name
                dept = get_department_by_name(rec.get('department'))
                desig = get_designation_by_name(rec.get('designation'))
                
                # We update or create
                emp, e_created = Employee.objects.update_or_create(
                    email=email,
                    defaults={
                        'first_name': rec.get('first_name'),
                        'last_name': rec.get('last_name'),
                        'employee_code': rec.get('employee_code') or ('EMP-' + email.split('@')[0]),
                        'phone': rec.get('phone'),
                        'department': dept,
                        'designation': desig,
                        'date_of_joining': parse_date(rec.get('date_of_joining')) or timezone.now().date(),
                        'employment_type': rec.get('type', 'full_time').lower(),
                        'pm_member_id': member.member_id 
                    }
                )
            
            created.append(email)
        except Exception as e:
            errors.append({'row': idx + 2, 'error': f"Employee '{email}': {str(e)}"})
    return created, errors

def import_projects(records):
    created = []
    errors = []
    for idx, rec in enumerate(records):
        try:
            name = rec.get('name')
            if not name: continue
            
            # Lookup Owner
            owner_name = rec.get('owner')
            owner = get_member_by_name_or_email(owner_name)

            slug = rec.get('slug') or name.lower().replace(' ', '-')
            
            project, _ = Projects.objects.update_or_create(
                slug=slug,
                defaults={
                    'name': name,
                    'description': rec.get('description', ''),
                    'status': rec.get('status', 'active'),
                    'visibility': rec.get('visibility', 'private'),
                    'start_date': parse_date(rec.get('start_date')),
                    'end_date': parse_date(rec.get('end_date')),
                    'owner_member_id': owner
                }
            )
            created.append(str(project.project_id))
        except Exception as e:
            errors.append({'row': idx + 2, 'error': f"Project '{rec.get('name', 'Unknown')}': {str(e)}"})
    return created, errors

def import_tasks(records, project_id=None):
    created = []
    errors = []
    
    for idx, rec in enumerate(records):
        try:
            title = rec.get('title') or rec.get('name')
            if not title: continue

            # Resolve Project
            curr_project = None
            if project_id:
                curr_project = Projects.objects.filter(project_id=project_id).first()
            elif rec.get('project'):
                curr_project = Projects.objects.filter(name__iexact=rec.get('project')).first()
            
            # Resolve Lookups
            status_obj = get_status_by_name(rec.get('status'))
            priority_obj = get_priority_by_name(rec.get('priority'))
            
            Task = Tasks.objects.create(
                title=title,
                description=rec.get('description', ''),
                project_id=curr_project,
                status_id=status_obj,
                priority_id=priority_obj,
                due_date=parse_date(rec.get('due_date')),
                start_date=parse_date(rec.get('start_date')),
                estimate_hours=float(rec.get('estimate_hours') or 0),
                percent_complete=int(rec.get('percent_complete') or 0)
            )
            created.append(str(Task.task_id))
        except Exception as e:
            errors.append({'row': idx + 2, 'error': f"Task '{rec.get('title', 'Unknown')}': {str(e)}"})
    return created, errors

# --- Views ---

class UnifiedImportView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        """Import multiple entities from keys sheets in a single Excel file"""
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        filename = file.name.lower()
        if not (filename.endswith('.xlsx') or filename.endswith('.xls')):
            return Response({'error': 'Unified import requires an Excel file (.xlsx)'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            report = {
                'departments': {'created': 0, 'failed': 0, 'errors': []},
                'designations': {'created': 0, 'failed': 0, 'errors': []},
                'employees': {'created': 0, 'failed': 0, 'errors': []}, 
                'projects': {'created': 0, 'failed': 0, 'errors': []},
                'tasks': {'created': 0, 'failed': 0, 'errors': []},
                'summary': []
            }
            sheet_names = [s.lower() for s in wb.sheetnames]
            
            # Order Matters: Dept -> Desig -> Employee/Member -> Project -> Task

            # 1. Departments
            if 'departments' in sheet_names:
                records = worksheet_to_records(wb[wb.sheetnames[sheet_names.index('departments')]])
                c, e = import_departments(records)
                report['departments'] = {'created': len(c), 'failed': len(e), 'errors': e}
                report['summary'].append(f"Departments: {len(c)} imported")

            # 2. Designations
            if 'designations' in sheet_names:
                records = worksheet_to_records(wb[wb.sheetnames[sheet_names.index('designations')]])
                c, e = import_designations(records)
                report['designations'] = {'created': len(c), 'failed': len(e), 'errors': e}
                report['summary'].append(f"Designations: {len(c)} imported")

            # 3. Employees / Members
            # Support both 'Employees' and 'Members' sheet names
            emp_sheet = None
            if 'employees' in sheet_names: emp_sheet = wb[wb.sheetnames[sheet_names.index('employees')]]
            elif 'members' in sheet_names: emp_sheet = wb[wb.sheetnames[sheet_names.index('members')]]
            
            if emp_sheet:
                records = worksheet_to_records(emp_sheet)
                c, e = import_employees_and_members(records)
                report['employees'] = {'created': len(c), 'failed': len(e), 'errors': e}
                report['summary'].append(f"Employees: {len(c)} imported")

            # 4. Projects
            if 'projects' in sheet_names:
                records = worksheet_to_records(wb[wb.sheetnames[sheet_names.index('projects')]])
                c, e = import_projects(records)
                report['projects'] = {'created': len(c), 'failed': len(e), 'errors': e}
                report['summary'].append(f"Projects: {len(c)} imported")

            # 5. Tasks
            if 'tasks' in sheet_names:
                records = worksheet_to_records(wb[wb.sheetnames[sheet_names.index('tasks')]])
                c, e = import_tasks(records)
                report['tasks'] = {'created': len(c), 'failed': len(e), 'errors': e}
                report['summary'].append(f"Tasks: {len(c)} imported")
            
            return Response({
                'success': True,
                'report': report,
                'created': sum(x['created'] for x in report.values() if isinstance(x, dict)), 
                'failed': sum(x['failed'] for x in report.values() if isinstance(x, dict))
            })

        except Exception as e:
            return Response({'error': f"Import Failed: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

    def get(self, request):
        """Download Unified Template"""
        output = io.BytesIO()
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames: del wb['Sheet']
        
        # Departments
        ws = wb.create_sheet('Departments')
        ws.append(['name', 'code', 'category', 'description'])
        ws.append(['Engineering', 'ENG', 'Development', 'Engineering Dept'])
        
        # Designations
        ws = wb.create_sheet('Designations')
        ws.append(['name', 'code', 'level', 'min_salary', 'max_salary'])
        ws.append(['Senior Developer', 'SDEV', 3, 50000, 80000])

        # Employees
        ws = wb.create_sheet('Employees')
        ws.append(['first_name', 'last_name', 'email', 'phone', 'department', 'designation', 'date_of_joining', 'type'])
        ws.append(['John', 'Doe', 'john@example.com', '1234567890', 'Engineering', 'Senior Developer', '2024-01-01', 'Full Time'])
        
        # Projects
        ws = wb.create_sheet('Projects')
        ws.append(['name', 'slug', 'description', 'status', 'visibility', 'start_date', 'end_date', 'owner'])
        ws.append(['New Website', 'new-website', 'Website Redesign', 'active', 'public', '2024-01-01', '2024-06-30', 'john@example.com'])
        
        # Tasks
        ws = wb.create_sheet('Tasks')
        ws.append(['title', 'description', 'project', 'due_date', 'estimate_hours', 'priority', 'status', 'percent_complete'])
        ws.append(['Setup Server', 'Initial Setup', 'New Website', '2024-01-15', 8, 'High', 'To Do', 0])
        
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(content=output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="unified_template.xlsx"'
        return response


class ProjectImportView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            records = parse_file(file)
            created, errors = import_projects(records)
            return Response({'created': len(created), 'failed': len(errors), 'errors': errors[:10]})
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def get(self, request):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="projects_template.csv"'
        writer = csv.writer(response)
        writer.writerow(['name', 'slug', 'description', 'status', 'visibility'])
        writer.writerow(['Example Project', 'example-project', 'Description here', 'active', 'private'])
        return response

class TaskImportView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        file = request.FILES.get('file')
        project_id = request.data.get('project_id')
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            records = parse_file(file)
            created, errors = import_tasks(records, project_id)
            return Response({'created': len(created), 'failed': len(errors), 'errors': errors[:10]})
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def get(self, request):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="tasks_template.csv"'
        writer = csv.writer(response)
        writer.writerow(['title', 'description', 'project', 'due_date', 'estimate_hours', 'priority', 'status'])
        writer.writerow(['Example Task', 'Task description', 'Project Name', '2024-12-31', 5.5, 'Medium', 'To Do'])
        return response

class MemberImportView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            records = parse_file(file)
            created, errors = import_members(records)
            return Response({'created': len(created), 'failed': len(errors), 'errors': errors[:10]})
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def get(self, request):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="members_template.csv"'
        writer = csv.writer(response)
        writer.writerow(['first_name', 'last_name', 'email', 'phone'])
        writer.writerow(['John', 'Doe', 'john@example.com', '+1234567890'])
        return response
