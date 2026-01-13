import requests
import openpyxl
import io
import os
import django

# Setup Django environment
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'backend.settings')
django.setup()

from pm.models import Projects, Tasks, Members

BASE_URL = 'http://127.0.0.1:8001/api/pm/import/unified/'

def create_unified_template():
    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames: del wb['Sheet']
    
    # 1. Departments
    ws = wb.create_sheet('Departments')
    ws.append(['name', 'code', 'category'])
    ws.append(['Engineering', 'ENG', 'Development'])
    
    # 2. Designations
    ws = wb.create_sheet('Designations')
    ws.append(['name', 'level', 'min_salary'])
    ws.append(['Senior Dev', 3, 60000])
    
    # 3. Employees
    ws = wb.create_sheet('Employees')
    ws.append(['first_name', 'last_name', 'email', 'phone', 'department', 'designation'])
    ws.append(['Full', 'Import', 'full@example.com', '1231231234', 'Engineering', 'Senior Dev'])
    
    # 4. Projects
    ws = wb.create_sheet('Projects')
    ws.append(['name', 'slug', 'description', 'status', 'owner', 'start_date'])
    ws.append(['Full Project', 'full-project', 'Desc', 'active', 'full@example.com', '2025-01-01'])
    
    # 5. Tasks
    ws = wb.create_sheet('Tasks')
    ws.append(['title', 'project', 'status', 'priority', 'percent_complete'])
    ws.append(['Full Task', 'Full Project', 'To Do', 'High', 50])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def verify_import():
    print("Testing Full Unified Import...")
    
    # Clean up
    Projects.objects.filter(slug='full-project').delete()
    Members.objects.filter(email='full@example.com').delete()
    try:
        from hr.models import Department
        Department.objects.filter(dept_name='Engineering').delete()
    except: pass
    
    file_content = create_unified_template()
    files = {'file': ('unified_full.xlsx', file_content, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
    
    try:
        response = requests.post(BASE_URL, files=files)
        print(f"Status: {response.status_code}")
        # print(response.json())
        if response.status_code == 200:
            rep = response.json().get('report', {})
            print("Report Summary:")
            for line in rep.get('summary', []):
                print(f" - {line}")
        else:
            print(f"Error: {response.text}")

    except Exception as e:
        print(f"Exception: {e}")

if __name__ == '__main__':
    verify_import()
